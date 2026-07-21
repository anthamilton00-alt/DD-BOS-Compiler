from openpyxl.styles import Alignment
from openpyxl.styles import Border
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from openpyxl.worksheet.table import Table
from openpyxl.worksheet.table import TableStyleInfo

from core.excel_formatter import format_workbook
HEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor="1F4E78",
)

HEADER_FONT = Font(
    bold=True,
    color="FFFFFF",
)

HEADER_ALIGNMENT = Alignment(
    horizontal="center",
    vertical="center",
)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

HIGH_RISK_FILL = PatternFill(
    fill_type="solid",
    fgColor="FFC7CE",
)

MEDIUM_RISK_FILL = PatternFill(
    fill_type="solid",
    fgColor="FFEB9C",
)


def format_workbook(workbook):
    """
    Apply styling and formatting to every worksheet.
    """

    table_index = 1

    for worksheet in workbook.worksheets:

        worksheet.freeze_panes = "A2"

        if worksheet.max_row < 1 or worksheet.max_column < 1:
            continue

        worksheet.auto_filter.ref = worksheet.dimensions

        for cell in worksheet[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = HEADER_ALIGNMENT
            cell.border = THIN_BORDER

        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.border = THIN_BORDER

        for column_cells in worksheet.columns:
            width = max(
                len(str(cell.value))
                if cell.value is not None
                else 0
                for cell in column_cells
            )

            worksheet.column_dimensions[
                column_cells[0].column_letter
            ].width = min(width + 3, 60)

        if worksheet.max_row > 1 and worksheet.max_column > 1:

            table = Table(
                displayName=f"Table{table_index}",
                ref=worksheet.dimensions,
            )

            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )

            worksheet.add_table(table)

            table_index += 1

    apply_risk_formatting(workbook)


def apply_risk_formatting(workbook):

    if "Risk Scores" not in workbook.sheetnames:
        return

    ws = workbook["Risk Scores"]

    headers = {
        cell.value: cell.column_letter
        for cell in ws[1]
    }

    if "Risk Score" not in headers:
        return

    column = headers["Risk Score"]

    high = DifferentialStyle(fill=HIGH_RISK_FILL)
    medium = DifferentialStyle(fill=MEDIUM_RISK_FILL)

    ws.conditional_formatting.add(
        f"{column}2:{column}{ws.max_row}",
        Rule(
            type="expression",
            formula=[f"{column}2>=10"],
            dxf=high,
        ),
    )

    ws.conditional_formatting.add(
        f"{column}2:{column}{ws.max_row}",
        Rule(
            type="expression",
            formula=[f"AND({column}2>=5,{column}2<10)"],
            dxf=medium,
        ),
    )    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, dataframe in sheets.items():
            dataframe.to_excel(
                writer,
                sheet_name=sheet_name,
                index=False,
            )

        format_workbook(writer.book)